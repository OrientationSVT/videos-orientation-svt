# -*- coding: utf-8 -*-
# Genere index.html a partir de l'Excel "Videos_orientation".
# Double-cliquez sur "lancer.bat" (ou executez : py generer_site.py).
import json, os, glob, sys

try:
    from openpyxl import load_workbook
except ImportError:
    print("La bibliotheque 'openpyxl' n'est pas installee.")
    print("Ouvrez une invite de commandes et tapez :  py -m pip install openpyxl")
    input("\nAppuyez sur Entree pour fermer...")
    sys.exit(1)


def trouver_excel(dossier):
    fichiers = [f for f in glob.glob(os.path.join(dossier, "*.xlsx"))
                if not os.path.basename(f).startswith("~$")]
    if not fichiers:
        return None
    prioritaires = [f for f in fichiers if "orientation" in os.path.basename(f).lower()]
    return max(prioritaires or fichiers, key=os.path.getmtime)


def _parse_codes(v):
    combos, libre = [], False
    if not v:
        return combos, libre
    for tok in str(v).split('|'):
        tok = tok.strip()
        if not tok:
            continue
        parts = [p.strip() for p in tok.split('+') if p.strip()]
        if '*' in parts:
            libre = True
            continue
        seen = []
        for p in (['SVT'] + [x for x in parts if x != 'SVT']):
            if p not in seen:
                seen.append(p)
        if len(seen) >= 2 and seen not in combos:
            combos.append(seen)
    return combos, libre


def extraire_donnees(chemin_excel):
    wb = load_workbook(chemin_excel)
    ws = wb['Présentations métiers']
    rg = wb['Regroupements']
    domaine_de = {}
    for r in range(2, rg.max_row + 1):
        dom = rg.cell(row=r, column=1).value
        cat = rg.cell(row=r, column=2).value
        if dom and cat:
            domaine_de[str(cat).strip()] = str(dom).strip()
    data, cur = [], None
    for r in range(5, ws.max_row + 1):
        name = ws.cell(row=r, column=2).value
        if not name:
            continue
        name = str(name).strip()
        c3 = ws.cell(row=r, column=3)
        c4 = ws.cell(row=r, column=4)
        specs = ws.cell(row=r, column=5).value
        opts = ws.cell(row=r, column=6).value
        desc = ws.cell(row=r, column=7).value
        deb = ws.cell(row=r, column=8).value
        code = ws.cell(row=r, column=9).value
        if (deb is not None) and (c3.hyperlink is None):
            combos, libre = _parse_codes(code)
            cur = {"name": name, "domain": domaine_de.get(name, "(non classé)"),
                   "specs": (specs or '').strip(), "options": (opts or '').strip(),
                   "debouches": (deb or '').strip(), "combos": combos,
                   "libreChoix": libre, "jobs": []}
            data.append(cur)
        else:
            u1 = c3.hyperlink.target if c3.hyperlink else None
            u2 = c4.hyperlink.target if c4.hyperlink else None
            if cur is not None and (u1 or u2 or desc):
                cur["jobs"].append({"name": name, "url1": u1, "url2": u2,
                                    "desc": (desc or '').strip()})
    for c in data:
        c["hasVideos"] = any(j["url1"] or j["url2"] for j in c["jobs"])
    return data


# Domaines : lus tels quels depuis le feuillet Regroupements (Excel fait foi).

DOMAINS = [
    "Sciences du vivant (biologie & génétique)",
    "Agriculture, alimentation & filières",
    "Environnement, biodiversité & écologie",
    "Sciences de la Terre, mer & univers",
    "Santé humaine & médecine",
    "Santé animale & vétérinaire",
    "Données, numérique & investigation",
    "Société, droit & entreprise",
    "Médiation, culture scientifique & enseignement",
    "Sport & activité physique",
]
DOMAIN_META = {
    "Sciences du vivant (biologie & génétique)": {"icon":"🧬","short":"Sciences du vivant"},
    "Agriculture, alimentation & filières": {"icon":"🌾","short":"Agriculture & alimentation"},
    "Environnement, biodiversité & écologie": {"icon":"🌍","short":"Environnement & écologie"},
    "Sciences de la Terre, mer & univers": {"icon":"🌊","short":"Terre, mer & univers"},
    "Santé humaine & médecine": {"icon":"🩺","short":"Santé humaine & médecine"},
    "Santé animale & vétérinaire": {"icon":"🐾","short":"Santé animale & vétérinaire"},
    "Données, numérique & investigation": {"icon":"💻","short":"Données & investigation"},
    "Société, droit & entreprise": {"icon":"⚖️","short":"Société, droit & entreprise"},
    "Médiation, culture scientifique & enseignement": {"icon":"📣","short":"Médiation & enseignement"},
    "Sport & activité physique": {"icon":"⚽","short":"Sport & activité physique"},
}
SPECIALTIES = [["PC","Physique-Chimie"],["Maths","Maths"],["HGGSP","HGGSP"],
    ["SES","SES"],["NSI","NSI"],["LLCER","LLCER"],["Arts","Arts"],["HLP","HLP"],["SI","SI"]]

ICONS = {"default":"🔬","Agronomie":"🌾","Agroalimentaire":"🥗","Biochimie":"⚗️","Biodiversité":"🦋","Bioinformatique":"💻","Biologie cellulaire":"🔬","Biologie moléculaire":"🧬","Biologie végétale":"🌿","Biotechnologies":"🧪","Commerce":"🤝","Data science":"📊","Droit":"⚖️","Epidémiologie":"📈","Ethologie et comportement animal":"🐾","Exobiologie":"🚀","Finance":"💹","Forêts":"🌲","Génétique":"🧬","Géosciences":"🪨","Géotechnique":"⛏️","Hydrogéologie":"💧","Météorologie":"🌤️","Missions scientifiques":"🧭","Muséographie":"🏛️","Neurosciences":"🧠","Océanographie":"🌊","Optique & Vision":"👁️","Paléontologie":"🦕","Pharmacie":"💊","Physiologie humaine":"❤️","Police scientifique":"🔍","Psychologie/Psychiatrie":"🧩","Responsabilité sociétale des entreprises":"♻️","Sciences de l'évolution":"🦎","Santé - Professions médicales et paramédicales":"🩺","Santé - Thérapies complémentaires, accompagnement et bien-être":"🌸","Sport":"⚽","Apiculture & Entomologie appliquée":"🐝","Aquaculture & Pêche durable":"🐟","Biomatériaux & Implants":"🦴","Biomimétisme & Éco-conception":"🌱","Cosmétique & Dermatologie":"✨","Enseignement & Recherche en SVT":"📚","Génie écologique & Restauration":"🌍","Jardins & Espaces verts":"🌻","Journalisme & Médiation scientifique":"📰","Muséologie scientifique":"🏺","Mycologie & Champignons":"🍄","One Health / Santé globale":"🌐","Toxicologie & Sécurité chimique":"⚠️","Vétérinaire & Santé animale":"🐾","Muséographie & Muséologie scientifique":"🏛️"}

TEMPLATE = r"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="description" content="Ressource pédagogique gratuite : découvrez des dizaines de métiers accessibles avec la spécialité SVT au lycée, à travers des témoignages vidéo courts de professionnels. Navigation par domaine et par combinaison de spécialités. Par Lucas Bollori, sous licence CC BY-NC-SA 4.0.">
<meta name="author" content="Lucas Bollori">
<meta property="og:title" content="Vidéos d'Orientation SVT">
<meta property="og:description" content="Des métiers à découvrir avec la spécialité SVT — explorez par domaine ou par combinaison de spécialités.">
<meta property="og:type" content="website">
<title>Vidéos d'Orientation SVT — Sciences du Vivant & de la Terre</title>
<link href="https://fonts.googleapis.com/css2?family=Playfair+Display:wght@600;700&family=Source+Sans+3:ital,wght@0,300;0,400;0,600;1,300&display=swap" rel="stylesheet">
<style>
:root{
  --g900:#123E25; --g800:#15532F; --g700:#1C6B3C; --g600:#218A48; --g500:#37A65C;
  --g300:#86C9A0; --g100:#E4F3E9; --g50:#F1F8F3;
  --e700:#8C3A1F; --e600:#AE4A2C; --e500:#C75A3B; --e100:#F7E9E2;
  --blue:#2F5E73;
  --ink:#16271D; --muted:#5A6B60; --paper:#F6FAF5; --card:#FFFFFF; --line:#DBE7DE;
  --shadow:0 2px 14px rgba(20,60,35,.08); --shadow-lg:0 10px 34px rgba(20,60,35,.14);
  --radius:12px; --radius-sm:7px;
}
*{box-sizing:border-box;margin:0;padding:0}
body{font-family:'Source Sans 3',sans-serif;background:var(--paper);color:var(--ink);min-height:100vh;font-weight:300;line-height:1.5}
a{color:inherit}
header{background:linear-gradient(135deg,var(--g900) 0%,var(--g700) 55%,var(--g500) 100%);color:#fff;position:relative;overflow:hidden}
header::before{content:'';position:absolute;inset:0;background:url("data:image/svg+xml,%3Csvg width='60' height='60' viewBox='0 0 60 60' xmlns='http://www.w3.org/2000/svg'%3E%3Cg fill='none' fill-rule='evenodd'%3E%3Cg fill='%23ffffff' fill-opacity='0.04'%3E%3Cpath d='M36 34v-4h-2v4h-4v2h4v4h2v-4h4v-2h-4zm0-30V0h-2v4h-4v2h4v4h2V6h4V4h-4zM6 34v-4H4v4H0v2h4v4h2v-4h4v-2H6zM6 4V0H4v4H0v2h4v4h2V6h4V4H6z'/%3E%3C/g%3E%3C/g%3E%3C/svg%3E")}
.header-inner{max-width:1100px;margin:0 auto;padding:46px 32px 38px;position:relative}
.header-badge{display:inline-block;background:rgba(255,255,255,.16);border:1px solid rgba(255,255,255,.28);padding:5px 14px;border-radius:30px;font-size:.78rem;letter-spacing:.06em;text-transform:uppercase;font-weight:600;margin-bottom:18px}
h1{font-family:'Playfair Display',serif;font-weight:700;font-size:2.7rem;line-height:1.08;margin-bottom:14px}
h1 span{color:var(--g300);display:inline-block;border-bottom:3px solid var(--e500);padding-bottom:2px}
.header-sub{font-size:1.05rem;max-width:640px;opacity:.94;font-weight:300}
.header-stats{display:flex;gap:34px;margin:26px 0 18px}
.stat{display:flex;flex-direction:column}
.stat-num{font-family:'Playfair Display',serif;font-size:2rem;font-weight:700;line-height:1}
.stat-label{font-size:.8rem;text-transform:uppercase;letter-spacing:.05em;opacity:.8;margin-top:3px}
.header-credit{font-size:.85rem;opacity:.85}
.header-credit a{color:#fff;text-decoration:underline;text-underline-offset:2px}
.header-note{font-size:.72rem;opacity:.68;margin-top:5px;letter-spacing:.03em}
.controls{position:sticky;top:0;z-index:30;background:rgba(246,250,245,.94);backdrop-filter:blur(8px);border-bottom:1px solid var(--line)}
.controls-inner{max-width:1100px;margin:0 auto;padding:12px 32px}
.controls-row{display:flex;gap:12px;align-items:center;flex-wrap:wrap}
.controls-row+.controls-row{margin-top:10px}
.search-wrap{flex:1;min-width:220px;display:flex;align-items:center;gap:9px;background:var(--card);border:1.5px solid var(--line);border-radius:30px;padding:9px 16px;color:var(--muted);transition:.18s}
.search-wrap:focus-within{border-color:var(--g500);box-shadow:0 0 0 3px rgba(55,166,92,.16)}
.search-wrap input{border:none;outline:none;flex:1;font-family:inherit;font-size:.96rem;color:var(--ink);background:transparent;font-weight:300}
.filter-btns{display:flex;gap:7px}
.filter-btn{border:1.5px solid var(--line);background:var(--card);color:var(--muted);padding:8px 15px;border-radius:30px;font-family:inherit;font-size:.88rem;font-weight:600;cursor:pointer;transition:.15s}
.filter-btn:hover{border-color:var(--g300);color:var(--g700)}
.filter-btn.active{background:var(--g700);border-color:var(--g700);color:#fff}
.result-count{font-size:.85rem;color:var(--muted);font-weight:600;margin-left:auto;text-align:right}
/* Sélecteur de spécialités */
.spec-wrap{position:relative}
.spec-btn{display:flex;align-items:center;gap:8px;border:1.5px solid var(--g500);background:var(--g50);color:var(--g800);padding:9px 16px;border-radius:30px;font-family:inherit;font-size:.9rem;font-weight:700;cursor:pointer;transition:.15s}
.spec-btn:hover{background:var(--g100)}
.spec-btn.set{background:var(--g700);border-color:var(--g700);color:#fff}
.spec-btn .caret{width:14px;height:14px;transition:transform .2s}
.spec-wrap.open .spec-btn .caret{transform:rotate(180deg)}
.spec-panel{position:absolute;top:calc(100% + 8px);left:0;z-index:50;width:min(430px,92vw);background:var(--card);border:1px solid var(--line);border-radius:var(--radius);box-shadow:var(--shadow-lg);padding:18px;display:none}
.spec-wrap.open .spec-panel{display:block}
.spec-panel h4{font-family:'Playfair Display',serif;font-weight:600;font-size:1.08rem;color:var(--g900);margin-bottom:5px}
.spec-help{font-size:.83rem;color:var(--muted);margin-bottom:13px;line-height:1.45}
.spec-chips{display:flex;flex-wrap:wrap;gap:7px;margin-bottom:13px}
.spec-chip{border:1.5px solid var(--line);background:#fff;color:var(--ink);padding:7px 13px;border-radius:30px;font-size:.86rem;font-weight:600;cursor:pointer;transition:.13s;font-family:inherit}
.spec-chip:hover{border-color:var(--g300)}
.spec-chip.on{background:var(--g600);border-color:var(--g600);color:#fff}
.spec-chip.dim{opacity:.38;pointer-events:none}
.spec-chip.locked{background:var(--g100);border-color:var(--g300);color:var(--g800);cursor:default;display:flex;align-items:center;gap:5px}
.spec-foot{display:flex;justify-content:space-between;align-items:center;font-size:.82rem;border-top:1px solid var(--line);padding-top:11px}
.spec-count{color:var(--g700);font-weight:700}
.spec-clear{background:none;border:none;color:var(--e600);font-weight:700;cursor:pointer;font-family:inherit;font-size:.82rem}
.spec-clear:hover{text-decoration:underline}
/* Navigation par domaine */
.domain-nav{display:flex;gap:8px;overflow-x:auto;padding:2px 0;scrollbar-width:thin;flex:1;min-width:0}
.domain-nav::-webkit-scrollbar{height:5px}
.domain-nav::-webkit-scrollbar-thumb{background:var(--g300);border-radius:5px}
.domain-chip{white-space:nowrap;border:1.5px solid var(--line);background:var(--card);color:var(--muted);padding:7px 13px;border-radius:30px;font-size:.84rem;font-weight:600;cursor:pointer;transition:.15s;flex-shrink:0;display:flex;align-items:center;gap:6px;font-family:inherit}
.domain-chip:hover{border-color:var(--g300);color:var(--g700)}
.domain-chip.active{background:var(--g100);border-color:var(--g500);color:var(--g800)}
.toc{margin:22px 0 4px}
.toc.hidden{display:none}
.toc-title{font-family:'Playfair Display',serif;font-weight:700;font-size:1.12rem;color:var(--g900);margin-bottom:13px;display:flex;align-items:center;gap:8px}
.toc-grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(215px,1fr));gap:10px}
.toc-card{display:flex;align-items:center;gap:12px;background:var(--card);border:1px solid var(--line);border-radius:var(--radius);padding:12px 14px;cursor:pointer;transition:.15s;text-align:left;font-family:inherit;width:100%}
.toc-card:hover{border-color:var(--g500);box-shadow:var(--shadow);transform:translateY(-1px)}
.toc-card:focus-visible{outline:3px solid var(--g500);outline-offset:2px}
.toc-icon{font-size:1.55rem;line-height:1;flex-shrink:0}
.toc-body{min-width:0;display:flex;flex-direction:column}
.toc-name{font-weight:700;font-size:.94rem;color:var(--ink);line-height:1.2}
.toc-count{font-size:.77rem;color:var(--muted);font-weight:600;margin-top:2px}
main{max-width:1100px;margin:0 auto;padding:20px 32px 60px}
.domain-section{margin:34px 0 8px;scroll-margin-top:132px}
.domain-section:first-child{margin-top:16px}
.domain-section.hidden{display:none}
.domain-header{display:flex;align-items:center;gap:12px;padding:8px 4px 12px;margin-bottom:14px;border-bottom:2px solid var(--g100)}
.domain-header .d-icon{font-size:1.5rem;line-height:1}
.domain-header .d-name{font-family:'Playfair Display',serif;font-weight:700;font-size:1.35rem;color:var(--g900)}
.domain-header .d-count{margin-left:auto;font-size:.78rem;font-weight:700;color:var(--g700);background:var(--g100);padding:4px 11px;border-radius:30px;white-space:nowrap}
.cat-card{background:var(--card);border:1px solid var(--line);border-radius:var(--radius);margin-bottom:14px;box-shadow:var(--shadow);overflow:hidden;transition:.2s}
.cat-card:hover{box-shadow:var(--shadow-lg)}
.cat-card.no-video{background:#FAFBFA}
.cat-card.hidden{display:none}
.cat-header{display:flex;justify-content:space-between;align-items:center;padding:16px 20px;cursor:pointer;gap:14px}
.cat-header:focus-visible{outline:3px solid var(--g500);outline-offset:-3px;border-radius:var(--radius)}
.cat-header-left{display:flex;align-items:center;gap:14px;min-width:0}
.cat-icon{width:42px;height:42px;flex-shrink:0;display:grid;place-items:center;background:var(--g100);border-radius:11px;font-size:1.3rem}
.no-video .cat-icon{background:var(--e100)}
.cat-name{font-family:'Playfair Display',serif;font-weight:600;font-size:1.16rem;color:var(--ink);min-width:0;overflow-wrap:anywhere}
.cat-meta{display:flex;align-items:center;gap:10px;flex-shrink:0}
.badge{font-size:.76rem;font-weight:700;padding:4px 11px;border-radius:30px;white-space:nowrap}
.badge-count{background:var(--g100);color:var(--g800)}
.badge-todo{background:var(--e100);color:var(--e700)}
.badge-match{background:var(--g600);color:#fff;display:flex;align-items:center;gap:4px}
.cat-toggle svg{width:20px;height:20px;color:var(--muted);transition:transform .22s}
.cat-card.open .cat-toggle svg{transform:rotate(90deg)}
.specs-bar{padding:0 20px 14px;display:flex;flex-direction:column;gap:7px}
.specs-row{display:flex;gap:10px;align-items:flex-start;font-size:.9rem}
.specs-chip{font-size:.68rem;font-weight:700;letter-spacing:.04em;text-transform:uppercase;padding:3px 9px;border-radius:6px;white-space:nowrap;flex-shrink:0;margin-top:1px}
.chip-spec{background:var(--g100);color:var(--g800)}
.chip-opt{background:#EAF1F4;color:var(--blue)}
.chip-deb{background:var(--e100);color:var(--e700)}
.specs-text{color:var(--ink);font-weight:400}
.specs-text.muted{color:var(--muted);font-weight:300}
.jobs-list{max-height:0;overflow:hidden;transition:max-height .3s ease}
.cat-card.open .jobs-list{max-height:8000px}
.job-row{display:flex;align-items:center;gap:14px;padding:11px 20px;border-top:1px solid var(--line)}
.job-row.odd{background:var(--g50)}
.no-video .job-row.odd{background:#F4F6F4}
.job-info{flex:1;min-width:0}
.job-name{font-size:.97rem;font-weight:600;color:var(--ink);display:block}
.job-name.placeholder{color:#9AA89E;font-style:italic;font-weight:400}
.job-desc{font-size:.84rem;color:var(--muted);font-weight:300;margin-top:2px}
.video-btn{display:inline-flex;align-items:center;gap:6px;padding:7px 13px;border-radius:30px;font-size:.83rem;font-weight:700;text-decoration:none;color:#fff;white-space:nowrap;transition:.15s;flex-shrink:0}
.video-btn svg{width:13px;height:13px}
.video-btn-1{background:linear-gradient(135deg,var(--g600),var(--g700))}
.video-btn-1:hover{filter:brightness(1.08);transform:translateY(-1px)}
.video-btn-2{background:linear-gradient(135deg,var(--e500),var(--e600))}
.video-btn-2:hover{filter:brightness(1.06);transform:translateY(-1px)}
.video-empty{width:1px}
.empty-state{display:none;text-align:center;padding:60px 20px;color:var(--muted)}
.empty-state svg{width:46px;height:46px;margin-bottom:14px;opacity:.5}
footer{background:var(--g900);color:#D9E7DD;text-align:center;padding:28px 32px;font-size:.86rem;line-height:1.7}
footer strong{color:#fff}
footer a{color:var(--g300);text-decoration:underline;text-underline-offset:2px}
@media(max-width:640px){
  h1{font-size:1.8rem}.header-inner{padding:24px 18px 20px}.header-stats{gap:20px;margin:16px 0 10px}
  .header-sub{font-size:.95rem}.stat-num{font-size:1.6rem}
  .controls-inner,main{padding-left:18px;padding-right:18px}
  .job-row{flex-wrap:wrap}.video-btn{font-size:.8rem;padding:6px 11px}
  .cat-name{font-size:1.05rem}.domain-header .d-name{font-size:1.15rem}
  .cat-header{flex-direction:column;align-items:flex-start;gap:9px}
  .toc-grid{grid-template-columns:1fr 1fr}
  .result-count{margin-left:0;width:100%}
}
@media(prefers-reduced-motion:reduce){*{transition:none!important;animation:none!important}}
</style>
</head>
<body>
<header>
  <div class="header-inner">
    <h1>Vidéos d'Orientation<br><span>Sciences du Vivant &amp; de la Terre</span></h1>
    <p class="header-sub">Découvrez des métiers accessibles avec la spécialité SVT, à travers des témoignages vidéo courts de professionnels. Explorez par domaine, ou indiquez vos spécialités pour voir les pistes qui vous correspondent.</p>
    <div class="header-stats">
      <div class="stat"><span class="stat-num" id="stat-cats">47</span><span class="stat-label">Catégories</span></div>
      <div class="stat"><span class="stat-num" id="stat-jobs">300</span><span class="stat-label">Métiers</span></div>
      <div class="stat"><span class="stat-num" id="stat-videos">316</span><span class="stat-label">Vidéos</span></div>
    </div>
    <p class="header-credit">Par Lucas Bollori, professeur de SVT &nbsp;·&nbsp; <a href="https://creativecommons.org/licenses/by-nc-sa/4.0/" target="_blank" rel="noopener noreferrer">CC BY-NC-SA 4.0</a></p>
    <p class="header-note">Ressource pédagogique en libre accès</p>
  </div>
</header>
<div class="controls">
  <div class="controls-inner">
    <div class="controls-row">
      <div class="search-wrap">
        <svg width="16" height="16" fill="none" stroke="currentColor" stroke-width="2" viewBox="0 0 24 24"><circle cx="11" cy="11" r="8"/><path d="m21 21-4.35-4.35"/></svg>
        <input type="text" id="search" placeholder="Rechercher un métier ou une catégorie…" autocomplete="off">
      </div>
      <div class="spec-wrap" id="spec-wrap">
        <button class="spec-btn" id="spec-btn" aria-expanded="false" aria-haspopup="true">
          <span>🎓 <span id="spec-btn-label">Mes spécialités</span></span>
          <svg class="caret" fill="none" stroke="currentColor" stroke-width="2.5" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" d="m6 9 6 6 6-6"/></svg>
        </button>
        <div class="spec-panel" id="spec-panel" role="dialog" aria-label="Choisir mes spécialités">
          <h4>Quelles sont tes spécialités&nbsp;?</h4>
          <p class="spec-help">Toutes ces pistes supposent que tu gardes la <strong>SVT</strong>. Ajoute la ou les spécialités que tu comptes prendre — <strong>2 en plus</strong> (profil Première) ou <strong>1 conservée</strong> (Terminale). Les catégories dont les spécialités recommandées correspondent s'afficheront.</p>
          <div class="spec-chips" id="spec-chips"></div>
          <div class="spec-foot">
            <span class="spec-count" id="spec-count">Aucune spécialité ajoutée</span>
            <button class="spec-clear" id="spec-clear">Effacer</button>
          </div>
        </div>
      </div>
    </div>
    <div class="controls-row">
      <div class="domain-nav" id="domain-nav"></div>
      <span class="result-count" id="result-count"></span>
    </div>
  </div>
</div>
<main>
  <div class="toc" id="toc"></div>
  <div id="domains"></div>
  <div class="empty-state" id="empty-state">
    <svg fill="none" stroke="currentColor" stroke-width="1.5" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" d="M21 21l-5.197-5.197m0 0A7.5 7.5 0 105.196 5.196a7.5 7.5 0 0010.607 10.607z"/></svg>
    <p>Aucun résultat pour cette sélection.</p>
  </div>
</main>
<footer>
  <strong>Vidéos d'Orientation SVT</strong> &nbsp;·&nbsp; Lucas Bollori, professeur de SVT<br>
  Licence <a href="https://creativecommons.org/licenses/by-nc-sa/4.0/" target="_blank" rel="noopener noreferrer">Creative Commons CC BY-NC-SA 4.0</a> &nbsp;·&nbsp; Usage pédagogique libre &nbsp;·&nbsp; Les vidéos sont la propriété de leurs auteurs respectifs
</footer>
<script>
const DATA = __DATA__;
const DOMAINS = __DOMAINS__;
const DOMAIN_META = __DOMAIN_META__;
const SPECIALTIES = __SPECIALTIES__;
const ICONS = __ICONS__;
const SPEC_LABEL = Object.fromEntries(SPECIALTIES.map(s=>[s[0],s[1]]));

let currentStatus="all", currentDomain="all", currentSearch="";
let selectedSpecs=new Set();

function getIcon(n){return ICONS[n]||ICONS.default}
function esc(s){return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;').replace(/"/g,'&quot;')}
function slug(s){return s.normalize('NFD').replace(/[\u0300-\u036f]/g,'').replace(/[^a-zA-Z0-9]+/g,'-').toLowerCase()}

function buildCard(cat){
  const hasVid=cat.hasVideos;
  const nbMet=cat.jobs.length;
  const nbVid=cat.jobs.reduce((s,j)=>s+(j.url1?1:0)+(j.url2?1:0),0);
  const jobRows=cat.jobs.map((j,i)=>{
    const placeholder=!j.url1&&!j.url2&&!hasVid;
    const desc=j.desc?`<span class="job-desc">${esc(j.desc)}</span>`:'';
    const info=`<div class="job-info"><span class="job-name${placeholder?' placeholder':''}">${esc(j.name)}</span>${desc}</div>`;
    const b1=j.url1?`<a class="video-btn video-btn-1" href="${esc(j.url1)}" target="_blank" rel="noopener noreferrer"><svg viewBox="0 0 24 24" fill="currentColor"><path d="M8 5v14l11-7z"/></svg> Vidéo 1</a>`:`<span class="video-empty"></span>`;
    const b2=j.url2?`<a class="video-btn video-btn-2" href="${esc(j.url2)}" target="_blank" rel="noopener noreferrer"><svg viewBox="0 0 24 24" fill="currentColor"><path d="M8 5v14l11-7z"/></svg> Vidéo 2</a>`:`<span class="video-empty"></span>`;
    return `<div class="job-row ${i%2===0?'odd':''}">${info}${b1}${b2}</div>`;
  }).join('');
  const badge=hasVid?`<span class="badge badge-count">${nbMet} métier${nbMet>1?'s':''} · ${nbVid} vidéo${nbVid>1?'s':''}</span>`:`<span class="badge badge-todo">À compléter</span>`;
  const optRow=cat.options?`<div class="specs-row"><span class="specs-chip chip-opt">Option</span><span class="specs-text muted">${esc(cat.options)}</span></div>`:'';
  const debRow=cat.debouches?`<div class="specs-row"><span class="specs-chip chip-deb">Études envisageables</span><span class="specs-text muted">${esc(cat.debouches)}</span></div>`:'';
  const partners=cat.combos.map(c=>c.filter(x=>x!=='SVT'));
  return `
<div class="cat-card${hasVid?'':' no-video'}" data-name="${esc(cat.name.toLowerCase())}" data-has-videos="${hasVid}" data-domain="${esc(cat.domain)}" data-libre="${!!cat.libreChoix}" data-combos='${JSON.stringify(partners)}'>
  <div class="cat-header" role="button" tabindex="0" aria-expanded="false">
    <div class="cat-header-left"><div class="cat-icon">${getIcon(cat.name)}</div><span class="cat-name">${esc(cat.name)}</span></div>
    <div class="cat-meta"><span class="badge badge-match" style="display:none">✓ Ta combinaison</span>${badge}<span class="cat-toggle"><svg fill="none" stroke="currentColor" stroke-width="2.5" viewBox="0 0 24 24"><path stroke-linecap="round" stroke-linejoin="round" d="m9 18 6-6-6-6"/></svg></span></div>
  </div>
  <div class="specs-bar">
    <div class="specs-row"><span class="specs-chip chip-spec">Spécialités</span><span class="specs-text">${esc(cat.specs)}</span></div>
    ${optRow}${debRow}
  </div>
  <div class="jobs-list">${jobRows}</div>
</div>`;
}

function sortCats(a,b){ if(a.hasVideos!==b.hasVideos) return a.hasVideos?-1:1; return a.name.localeCompare(b.name,'fr'); }

function render(){
  let html='';
  DOMAINS.forEach(dom=>{
    const cats=DATA.filter(c=>c.domain===dom).sort(sortCats);
    if(!cats.length) return;
    const m=DOMAIN_META[dom]||{icon:'🔬',short:dom};
    const domMet=cats.reduce((a,c)=>a+c.jobs.length,0);
    html+=`<section class="domain-section" data-domain="${esc(dom)}" id="dom-${slug(dom)}">
      <div class="domain-header"><span class="d-icon">${m.icon}</span><span class="d-name">${esc(m.short)}</span><span class="d-count">${cats.length} catégorie${cats.length>1?'s':''} · ${domMet} métier${domMet>1?'s':''}</span></div>
      <div class="domain-cats">${cats.map(buildCard).join('')}</div>
    </section>`;
  });
  document.getElementById('domains').innerHTML=html;

  document.getElementById('stat-cats').textContent=DATA.length;
  document.getElementById('stat-jobs').textContent=DATA.reduce((a,c)=>a+c.jobs.length,0);
  document.getElementById('stat-videos').textContent=DATA.reduce((a,c)=>a+c.jobs.reduce((s,j)=>s+(j.url1?1:0)+(j.url2?1:0),0),0);

  buildDomainNav();
  buildToc();
  buildSpecChips();
  applyFilters();
}

function buildDomainNav(){
  const nav=document.getElementById('domain-nav');
  let h=`<button class="domain-chip active" data-domain="all">Tous les domaines</button>`;
  DOMAINS.forEach(dom=>{
    if(!DATA.some(c=>c.domain===dom)) return;
    const m=DOMAIN_META[dom];
    h+=`<button class="domain-chip" data-domain="${esc(dom)}">${m.icon} ${esc(m.short)}</button>`;
  });
  nav.innerHTML=h;
  nav.querySelectorAll('.domain-chip').forEach(btn=>btn.addEventListener('click',()=>{
    currentDomain=btn.dataset.domain;
    nav.querySelectorAll('.domain-chip').forEach(x=>x.classList.remove('active'));
    btn.classList.add('active');
    applyFilters();
    if(currentDomain==='all'){ const reduce=window.matchMedia('(prefers-reduced-motion: reduce)').matches; window.scrollTo({top:0,behavior:reduce?'auto':'smooth'}); }
  }));
}

function buildToc(){
  const el=document.getElementById('toc');
  let h=`<div class="toc-title">🧭 Parcourir par domaine</div><div class="toc-grid">`;
  DOMAINS.forEach(dom=>{
    const cats=DATA.filter(c=>c.domain===dom); if(!cats.length) return;
    const met=cats.reduce((a,c)=>a+c.jobs.length,0);
    const m=DOMAIN_META[dom]||{icon:'🔬',short:dom};
    h+=`<button class="toc-card" data-target="dom-${slug(dom)}"><span class="toc-icon">${m.icon}</span><span class="toc-body"><span class="toc-name">${esc(m.short)}</span><span class="toc-count">${cats.length} catégorie${cats.length>1?'s':''} · ${met} métier${met>1?'s':''}</span></span></button>`;
  });
  h+=`</div>`;
  el.innerHTML=h;
  el.querySelectorAll('.toc-card').forEach(b=>b.addEventListener('click',()=>{
    const sec=document.getElementById(b.dataset.target);
    const reduce=window.matchMedia('(prefers-reduced-motion: reduce)').matches;
    if(sec) sec.scrollIntoView({behavior:reduce?'auto':'smooth',block:'start'});
  }));
}

function buildSpecChips(){
  const box=document.getElementById('spec-chips');
  let h=`<span class="spec-chip locked">✓ SVT</span>`;
  SPECIALTIES.forEach(([code,label])=>{ h+=`<button type="button" class="spec-chip" data-code="${code}">${esc(label)}</button>`; });
  box.innerHTML=h;
  box.querySelectorAll('button[data-code]').forEach(b=>b.addEventListener('click',()=>toggleSpec(b.dataset.code)));
  refreshSpecChips();
}
function toggleSpec(code){
  if(selectedSpecs.has(code)) selectedSpecs.delete(code);
  else{ if(selectedSpecs.size>=2) return; selectedSpecs.add(code); }
  refreshSpecChips(); updateSpecSummary(); applyFilters();
}
function refreshSpecChips(){
  document.querySelectorAll('#spec-chips button[data-code]').forEach(b=>{
    const on=selectedSpecs.has(b.dataset.code);
    b.classList.toggle('on',on);
    b.classList.toggle('dim', selectedSpecs.size>=2 && !on);
  });
  const n=selectedSpecs.size;
  document.getElementById('spec-count').textContent = n===0?'Aucune spécialité ajoutée'
    : n+' ajoutée'+(n>1?'s':'')+(n===2?' · profil Première':' · profil Terminale');
}
function updateSpecSummary(){
  const lbl=document.getElementById('spec-btn-label'), btn=document.getElementById('spec-btn');
  if(selectedSpecs.size===0){ lbl.textContent='Mes spécialités'; btn.classList.remove('set'); }
  else{ lbl.textContent='SVT + '+[...selectedSpecs].map(c=>SPEC_LABEL[c]||c).join(' + '); btn.classList.add('set'); }
}
function clearSpecs(){ selectedSpecs.clear(); refreshSpecChips(); updateSpecSummary(); applyFilters(); }

function applyFilters(){
  const q=currentSearch.toLowerCase().trim();
  const specActive=selectedSpecs.size>0;
  let visible=0; const domCount={};
  document.querySelectorAll('.cat-card').forEach(card=>{
    const name=card.dataset.name, hasVideos=card.dataset.hasVideos==='true';
    const cdom=card.dataset.domain, libre=card.dataset.libre==='true';
    let combos=[]; try{combos=JSON.parse(card.dataset.combos);}catch(e){}
    const txt=[...card.querySelectorAll('.job-name,.job-desc')].map(e=>e.textContent.toLowerCase());
    const matchSearch=!q||name.includes(q)||txt.some(t=>t.includes(q));
    const matchStatus=currentStatus==='all'||(currentStatus==='videos'&&hasVideos)||(currentStatus==='todo'&&!hasVideos);
    const matchDomain=currentDomain==='all'||cdom===currentDomain;
    const matchSpec=!specActive||libre||combos.some(ps=>ps.length>0&&ps.every(p=>selectedSpecs.has(p)));
    const show=matchSearch&&matchStatus&&matchDomain&&matchSpec;
    card.classList.toggle('hidden',!show);
    const mb=card.querySelector('.badge-match');
    if(mb) mb.style.display=(specActive&&show)?'flex':'none';
    if(show){
      visible++; domCount[cdom]=(domCount[cdom]||0)+1;
      if(q){card.classList.add('open');card.querySelector('.cat-header').setAttribute('aria-expanded','true');
        card.querySelectorAll('.job-row').forEach(row=>{const n=row.querySelector('.job-name');const d=row.querySelector('.job-desc');const hit=(n&&n.textContent.toLowerCase().includes(q))||(d&&d.textContent.toLowerCase().includes(q));row.style.opacity=hit?'1':'0.4';});
      }else{card.querySelectorAll('.job-row').forEach(row=>row.style.opacity='1');}
    }
  });
  document.querySelectorAll('.domain-section').forEach(sec=>{
    sec.classList.toggle('hidden', !(domCount[sec.dataset.domain]>0));
  });
  const showToc = currentDomain==='all' && !q && !specActive && currentStatus==='all';
  document.getElementById('toc').classList.toggle('hidden', !showToc);
  const rc=document.getElementById('result-count');
  if(specActive){ rc.textContent=`${visible} catégorie${visible>1?'s':''} pour SVT + ${[...selectedSpecs].map(c=>SPEC_LABEL[c]||c).join(' + ')}`; }
  else if(q){ rc.textContent=`${visible} catégorie${visible>1?'s':''} trouvée${visible>1?'s':''}`; }
  else if(currentDomain!=='all'||currentStatus!=='all'){ rc.textContent=`${visible} catégorie${visible>1?'s':''}`; }
  else rc.textContent='';
  document.getElementById('empty-state').style.display=visible===0?'block':'none';
}

// Événements
let st;
document.getElementById('search').addEventListener('input',e=>{clearTimeout(st);st=setTimeout(()=>{currentSearch=e.target.value;applyFilters();},200);});
document.querySelectorAll('.filter-btn').forEach(b=>b.addEventListener('click',()=>{
  document.querySelectorAll('.filter-btn').forEach(x=>x.classList.remove('active'));
  b.classList.add('active'); currentStatus=b.dataset.filter; applyFilters();
}));
const specWrap=document.getElementById('spec-wrap');
document.getElementById('spec-btn').addEventListener('click',e=>{
  e.stopPropagation(); const open=specWrap.classList.toggle('open');
  document.getElementById('spec-btn').setAttribute('aria-expanded',open);
});
document.getElementById('spec-panel').addEventListener('click',e=>e.stopPropagation());
document.getElementById('spec-clear').addEventListener('click',clearSpecs);
document.addEventListener('click',()=>{ if(specWrap.classList.contains('open')){specWrap.classList.remove('open');document.getElementById('spec-btn').setAttribute('aria-expanded','false');} });
document.addEventListener('keydown',e=>{ if(e.key==='Escape'&&specWrap.classList.contains('open')){specWrap.classList.remove('open');document.getElementById('spec-btn').setAttribute('aria-expanded','false');} });

// Dépliage des cartes (clic + clavier)
const domainsEl=document.getElementById('domains');
domainsEl.addEventListener('click',e=>{
  if(e.target.closest('a')) return;
  const h=e.target.closest('.cat-header'); if(!h) return;
  const open=h.parentElement.classList.toggle('open');
  h.setAttribute('aria-expanded',open);
});
domainsEl.addEventListener('keydown',e=>{
  const h=e.target.closest('.cat-header'); if(!h) return;
  if(e.key==='Enter'||e.key===' '){ e.preventDefault(); const open=h.parentElement.classList.toggle('open'); h.setAttribute('aria-expanded',open); }
});

render();
</script>
</body>
</html>"""

def main():
    dossier = os.path.dirname(os.path.abspath(__file__))
    excel = trouver_excel(dossier)
    if not excel:
        print("Aucun fichier Excel (.xlsx) trouve dans ce dossier :")
        print("  ", dossier)
        print("Placez ce script dans le meme dossier que votre Excel, puis relancez.")
        input("\nAppuyez sur Entree pour fermer...")
        return
    print("Fichier Excel utilise :", os.path.basename(excel))
    data = extraire_donnees(excel)
    out = (TEMPLATE
        .replace("__DATA__", json.dumps(data, ensure_ascii=False))
        .replace("__DOMAINS__", json.dumps(DOMAINS, ensure_ascii=False))
        .replace("__DOMAIN_META__", json.dumps(DOMAIN_META, ensure_ascii=False))
        .replace("__SPECIALTIES__", json.dumps(SPECIALTIES, ensure_ascii=False))
        .replace("__ICONS__", json.dumps(ICONS, ensure_ascii=False)))
    sortie = os.path.join(dossier, "index.html")
    with open(sortie, "w", encoding="utf-8") as f:
        f.write(out)
    nb_cat = len(data)
    nb_met = sum(len(c["jobs"]) for c in data)
    nb_vid = sum((1 if j["url1"] else 0) + (1 if j["url2"] else 0)
                 for c in data for j in c["jobs"])
    non_classe = [c["name"] for c in data if c["domain"] == "(non classé)"]
    print("index.html regenere avec succes.")
    print("  %d categories - %d metiers - %d videos" % (nb_cat, nb_met, nb_vid))
    if non_classe:
        print("  /!\\ Categories sans domaine (a ajouter au feuillet Regroupements) :")
        for n in non_classe:
            print("     -", n)
    print("\nVous pouvez maintenant publier avec GitHub Desktop.")
    input("\nAppuyez sur Entree pour fermer...")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print("Erreur inattendue :", e)
        input("\nAppuyez sur Entree pour fermer...")
